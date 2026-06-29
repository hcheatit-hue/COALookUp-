from flask import Flask, jsonify, request, render_template, send_file
import pyodbc
import os
import logging
from dotenv import load_dotenv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load environment variables
load_dotenv()

app = Flask(__name__)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# --- Flask Configuration ---
FLASK_HOST = os.getenv('FLASK_HOST', '0.0.0.0')
FLASK_PORT = int(os.getenv('FLASK_PORT', 5055))
FLASK_DEBUG = os.getenv('FLASK_DEBUG', 'False').lower() == 'true'

# --- DB Configuration ---
DB_SERVER = os.getenv('DB_SERVER', 'HCHDB01\\SSINFOR')
DB_NAME = os.getenv('DB_NAME', 'SSHCHEAT')
DB_DRIVER = os.getenv('DB_DRIVER', '{ODBC Driver 17 for SQL Server}')
DB_USER = os.getenv('DB_USER', '')
DB_PASSWORD = os.getenv('DB_PASSWORD', '')
DB_ENCRYPT = os.getenv('DB_ENCRYPT', 'yes')
DB_TRUST_CERT = os.getenv('DB_TRUST_CERT', 'yes')

def get_conn():
    """Get database connection with error handling."""
    try:
        # Build connection string based on available credentials
        if DB_USER and DB_PASSWORD:
            # SQL Server authentication
            conn_string = (
                f"DRIVER={DB_DRIVER};"
                f"SERVER={DB_SERVER};"
                f"DATABASE={DB_NAME};"
                f"UID={DB_USER};"
                f"PWD={DB_PASSWORD};"
                f"Encrypt={DB_ENCRYPT};"
                f"TrustServerCertificate={DB_TRUST_CERT};"
            )
            logger.debug(f"Using SQL Server authentication for {DB_USER}@{DB_SERVER}")
        else:
            # Windows integrated authentication
            conn_string = (
                f"DRIVER={DB_DRIVER};"
                f"SERVER={DB_SERVER};"
                f"DATABASE={DB_NAME};"
                f"Trusted_Connection=yes;"
                f"Encrypt={DB_ENCRYPT};"
                f"TrustServerCertificate={DB_TRUST_CERT};"
            )
            logger.debug(f"Using Windows integrated authentication for {DB_SERVER}")
        
        conn = pyodbc.connect(conn_string, autocommit=True)
        logger.info(f"Database connection successful to {DB_SERVER}/{DB_NAME}")
        return conn
    except pyodbc.DatabaseError as e:
        logger.error(f"Database connection failed: {str(e)}")
        raise
    except Exception as e:
        logger.error(f"Unexpected connection error: {str(e)}")
        raise

# --- Status code mapping ---
STATUS_MAP = {
    'O': {'label': 'Open',       'color': 'status-open'},
    'B': {'label': 'Backordered','color': 'status-backorder'},
    'X': {'label': 'Closed',     'color': 'status-closed'},
    'C': {'label': 'Cancelled',  'color': 'status-cancelled'},
    'P': {'label': 'Picked',     'color': 'status-picked'},
    'S': {'label': 'Shipped',    'color': 'status-shipped'},
    'I': {'label': 'Invoiced',   'color': 'status-invoiced'},
    'H': {'label': 'On Hold',    'color': 'status-hold'},
    'Q': {'label': 'Quote',      'color': 'status-quote'},
}

LINE_STATUS_MAP = {
    'O': {'label': 'Open',      'color': 'status-open'},
    'X': {'label': 'Closed',    'color': 'status-closed'},
    'C': {'label': 'Cancelled', 'color': 'status-cancelled'},
    'S': {'label': 'Shipped',   'color': 'status-shipped'},
    'B': {'label': 'Backordered','color': 'status-backorder'},
    'A': {'label': 'Active',    'color': 'status-open'},
    'H': {'label': 'On Hold',   'color': 'status-hold'},
}

# Add Complete status for 100% fulfilled orders
STATUS_MAP['COMPLETE'] = {'label': 'Complete', 'color': 'status-complete'}

def is_order_complete(order_lines):
    """Check if an order is complete by verifying all line items are fully shipped."""
    if not order_lines:
        return False
    
    for line in order_lines:
        order_qty = float(line.get('order_qty', 0))
        shipped_qty = float(line.get('shipped_qty', 0))
        
        # If any line has shipped quantity less than ordered quantity, order is not complete
        if shipped_qty < order_qty:
            return False
    
    return True

def fmt_date(val):
    if val is None:
        return None
    try:
        return val.strftime('%d %b %Y')
    except:
        return str(val)

def fmt_currency(val, currency='SAR'):
    if val is None:
        return None
    symbol = 'R' if currency in ('SAR', 'ZAR') else currency + ' '
    return f"{symbol}{float(val):,.2f}"

def row_to_dict(row, cursor):
    cols = [col[0] for col in cursor.description]
    return dict(zip(cols, row))

# --- GL Specific Constants and Functions ---
PERIOD_TO_MONTH = {1:3,2:4,3:5,4:6,5:7,6:8,7:9,8:10,9:11,10:12,11:1,12:2}

def build_gl_query(search, start_year, start_period, end_year, end_period):
    sql = """
    SELECT
        a.ID                        AS Account_ID,
        a.DESCRIPTION               AS Description,
        ab.ACCT_YEAR                AS Fiscal_Year,
        ab.ACCT_PERIOD              AS Fiscal_Period,
        FORMAT(DATEFROMPARTS(
            CASE WHEN ab.ACCT_PERIOD >= 11 THEN ab.ACCT_YEAR ELSE ab.ACCT_YEAR - 1 END,
            CASE ab.ACCT_PERIOD
                WHEN 1 THEN 3 WHEN 2 THEN 4 WHEN 3 THEN 5 WHEN 4 THEN 6
                WHEN 5 THEN 7 WHEN 6 THEN 8 WHEN 7 THEN 9 WHEN 8 THEN 10
                WHEN 9 THEN 11 WHEN 10 THEN 12 WHEN 11 THEN 1 WHEN 12 THEN 2
            END, 1), 'MMM/yyyy')    AS Period,
        ab.CURR_BALANCE             AS Balance,
        (ab.DEBIT_BUDGET - ab.CREDIT_BUDGET) AS Budget,
        SUM(ab.DEBIT_BUDGET - ab.CREDIT_BUDGET)
            OVER (PARTITION BY ab.ACCOUNT_ID, ab.ACCT_YEAR ORDER BY ab.ACCT_PERIOD)
                                    AS Cumulative_Budget,
        ab.CURR_BALANCE - (ab.DEBIT_BUDGET - ab.CREDIT_BUDGET) AS Variance
    FROM ACCOUNT a
    INNER JOIN ACCOUNT_BALANCE ab ON a.ID = ab.ACCOUNT_ID
    WHERE (a.ID LIKE ? OR a.DESCRIPTION LIKE ?)
      AND (
            (ab.ACCT_YEAR > ? OR (ab.ACCT_YEAR = ? AND ab.ACCT_PERIOD >= ?))
        AND (ab.ACCT_YEAR < ? OR (ab.ACCT_YEAR = ? AND ab.ACCT_PERIOD <= ?))
      )
    ORDER BY a.ID, ab.ACCT_YEAR, ab.ACCT_PERIOD
    """
    like = f'%{search}%'
    params = (
        like, like,
        start_year, start_year, start_period,
        end_year,   end_year,   end_period
    )
    return sql, params

# --- ROUTES ---

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/health')
def health():
    """Health check endpoint - verifies database connectivity."""
    try:
        conn = get_conn()
        conn.close()
        auth_method = 'SQL Server' if DB_USER else 'Windows Integrated'
        return jsonify({
            'status': 'healthy',
            'database': DB_NAME,
            'server': DB_SERVER,
            'auth': auth_method,
            'port': FLASK_PORT
        }), 200
    except Exception as e:
        logger.error(f"Health check failed: {str(e)}")
        auth_method = 'SQL Server' if DB_USER else 'Windows Integrated'
        return jsonify({
            'status': 'unhealthy',
            'error': str(e),
            'database': DB_NAME,
            'server': DB_SERVER,
            'auth': auth_method,
            'port': FLASK_PORT
        }), 503

@app.route('/api/customer/search')
def customer_search():
    try:
        q = request.args.get('q', '').strip()
        if len(q) < 2:
            return jsonify([])
        
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT TOP 15
                ID,
                NAME,
                CITY,
                SALESREP_ID,
                CURRENCY_ID,
                ACTIVE_FLAG,
                LAST_ORDER_DATE
            FROM CUSTOMER
            WHERE NAME LIKE ? OR ID LIKE ?
            ORDER BY NAME
        """, (f'%{q}%', f'%{q}%'))
        rows = cur.fetchall()
        conn.close()
        
        results = []
        for r in rows:
            results.append({
                'id': r[0],
                'name': r[1],
                'city': r[2],
                'salesrep': r[3],
                'currency': r[4],
                'active': r[5],
                'last_order_date': fmt_date(r[6]),
            })
        return jsonify(results)
    except pyodbc.DatabaseError as e:
        logger.error(f"Database error in customer_search: {str(e)}")
        return jsonify({'error': 'Database connection failed'}), 503
    except Exception as e:
        logger.error(f"Error in customer_search: {str(e)}")
        return jsonify({'error': 'An unexpected error occurred'}), 500

@app.route('/api/customer/<customer_id>/orders')
def customer_orders(customer_id):
    try:
        # Get filter parameters
        status_filter = request.args.get('status', '').split(',') if request.args.get('status') else []
        order_date_from = request.args.get('order_date_from')
        order_date_to = request.args.get('order_date_to')
        desired_ship_date_from = request.args.get('desired_ship_date_from')
        desired_ship_date_to = request.args.get('desired_ship_date_to')
        salesrep = request.args.get('salesrep')
        territory = request.args.get('territory')
        currency = request.args.get('currency')
        total_ordered_min = request.args.get('total_ordered_min')
        total_ordered_max = request.args.get('total_ordered_max')
        order_type = request.args.get('order_type')
        backorder = request.args.get('backorder')
        limit = int(request.args.get('limit', 50))
        offset = int(request.args.get('offset', 0))

        conn = get_conn()
        cur = conn.cursor()

        # Build WHERE clause dynamically
        where_clauses = ["co.CUSTOMER_ID = ?"]
        params = [customer_id]

        if status_filter:
            placeholders = ','.join('?' for _ in status_filter)
            where_clauses.append(f"co.STATUS IN ({placeholders})")
            params.extend(status_filter)

        if order_date_from and order_date_to:
            where_clauses.append("co.ORDER_DATE BETWEEN ? AND ?")
            params.extend([order_date_from, order_date_to])
        elif order_date_from:
            where_clauses.append("co.ORDER_DATE >= ?")
            params.append(order_date_from)
        elif order_date_to:
            where_clauses.append("co.ORDER_DATE <= ?")
            params.append(order_date_to)

        if desired_ship_date_from and desired_ship_date_to:
            where_clauses.append("co.DESIRED_SHIP_DATE BETWEEN ? AND ?")
            params.extend([desired_ship_date_from, desired_ship_date_to])
        elif desired_ship_date_from:
            where_clauses.append("co.DESIRED_SHIP_DATE >= ?")
            params.append(desired_ship_date_from)
        elif desired_ship_date_to:
            where_clauses.append("co.DESIRED_SHIP_DATE <= ?")
            params.append(desired_ship_date_to)

        if salesrep:
            where_clauses.append("co.SALESREP_ID = ?")
            params.append(salesrep)

        if territory:
            where_clauses.append("co.TERRITORY LIKE ?")
            params.append(f'%{territory}%')

        if currency:
            where_clauses.append("co.CURRENCY_ID = ?")
            params.append(currency)

        if total_ordered_min and total_ordered_max:
            where_clauses.append("co.TOTAL_AMT_ORDERED BETWEEN ? AND ?")
            params.extend([float(total_ordered_min), float(total_ordered_max)])
        elif total_ordered_min:
            where_clauses.append("co.TOTAL_AMT_ORDERED >= ?")
            params.append(float(total_ordered_min))
        elif total_ordered_max:
            where_clauses.append("co.TOTAL_AMT_ORDERED <= ?")
            params.append(float(total_ordered_max))

        if order_type:
            where_clauses.append("co.ORDER_TYPE = ?")
            params.append(order_type)

        if backorder:
            where_clauses.append("co.BACKORDER_FLAG = ?")
            params.append(backorder)

        where_sql = " AND ".join(where_clauses)

        cur.execute(f"""
            SELECT
                co.ID,
                co.CUSTOMER_PO_REF,
                co.ORDER_DATE,
                co.DESIRED_SHIP_DATE,
                co.PROMISE_DATE,
                co.LAST_SHIPPED_DATE,
                co.STATUS,
                co.TOTAL_AMT_ORDERED,
                co.TOTAL_AMT_SHIPPED,
                co.CURRENCY_ID,
                co.SALESREP_ID,
                co.TERRITORY,
                co.USER_1,
                CASE WHEN EXISTS (
                    SELECT 1 FROM CUST_ORDER_LINE col 
                    WHERE (col.CUST_ORDER_ID = co.ID OR col.CUST_ORDER_ID = '`' + co.ID)
                    AND col.ORDER_QTY > ISNULL(col.TOTAL_SHIPPED_QTY, 0)
                ) THEN 0 ELSE 1 END AS IS_COMPLETE
            FROM CUSTOMER_ORDER co
            WHERE {where_sql}
            ORDER BY co.ORDER_DATE DESC
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
        """, params + [offset, limit])
        rows = cur.fetchall()
        conn.close()
        
        orders = []
        for r in rows:
            status_code = (r[6] or '').strip()
            is_complete = r[13]  # IS_COMPLETE field from the query
            
            # Override status to Complete if order is 100% fulfilled
            if is_complete:
                status_code = 'COMPLETE'
                status_info = STATUS_MAP['COMPLETE']
            else:
                status_info = STATUS_MAP.get(status_code, {'label': status_code or 'Unknown', 'color': 'status-unknown'})
            
            currency_val = r[9] or 'SAR'
            orders.append({
                'id': (r[0] or '').strip().lstrip('`'),
                'customer_po_ref': r[1],
                'order_date': fmt_date(r[2]),
                'desired_ship_date': fmt_date(r[3]),
                'promise_date': fmt_date(r[4]),
                'last_shipped_date': fmt_date(r[5]),
                'status_code': status_code,
                'status_label': status_info['label'],
                'status_color': status_info['color'],
                'total_ordered': fmt_currency(r[7], currency_val),
                'total_shipped': fmt_currency(r[8], currency_val),
                'total_ordered_raw': float(r[7] or 0),
                'total_shipped_raw': float(r[8] or 0),
                'currency': currency_val,
                'salesrep': r[10],
                'territory': r[11],
                'site': r[12],
            })
        return jsonify(orders)
    except pyodbc.DatabaseError as e:
        logger.error(f"Database error in customer_orders: {str(e)}")
        return jsonify({'error': 'Database connection failed'}), 503
    except Exception as e:
        logger.error(f"Error in customer_orders: {str(e)}")
        return jsonify({'error': 'An unexpected error occurred'}), 500

@app.route('/api/order/<order_id>')
def order_detail(order_id):
    try:
        # Get filter parameters for line items
        part_id = request.args.get('part_id')
        line_status_filter = request.args.get('line_status', '').split(',') if request.args.get('line_status') else []
        limit = int(request.args.get('limit', 50))
        offset = int(request.args.get('offset', 0))

        conn = get_conn()
        cur = conn.cursor()

        # --- Header ---
        cur.execute("""
            SELECT
                co.ID, co.CUSTOMER_ID, co.CUSTOMER_PO_REF,
                co.CONTACT_FIRST_NAME, co.CONTACT_LAST_NAME,
                co.CONTACT_PHONE, co.CONTACT_MOBILE, co.CONTACT_EMAIL,
                co.CONTACT_POSITION,
                co.ORDER_DATE, co.DESIRED_SHIP_DATE, co.PROMISE_DATE,
                co.LAST_SHIPPED_DATE, co.CREATE_DATE, co.STATUS_EFF_DATE,
                co.STATUS,
                co.TOTAL_AMT_ORDERED, co.TOTAL_AMT_SHIPPED,
                co.CURRENCY_ID, co.SELL_RATE, co.BUY_RATE,
                co.TERMS_DESCRIPTION, co.TERMS_NET_DAYS, co.TERMS_DISC_PERCENT,
                co.SALESREP_ID, co.TERRITORY,
                co.SHIP_VIA, co.FREE_ON_BOARD, co.FREIGHT_TERMS,
                co.BACK_ORDER, co.BACKORDER_FLAG,
                co.WAREHOUSE_ID, co.SITE_ID,
                co.USER_1, co.USER_2, co.USER_3,
                co.ENTERED_BY,
                c.NAME AS CUSTOMER_NAME,
                c.ADDR_1, c.ADDR_2, c.CITY, c.COUNTRY,
                co.TYPE, co.ORDER_TYPE
            FROM CUSTOMER_ORDER co
            LEFT JOIN CUSTOMER c ON co.CUSTOMER_ID = c.ID
            WHERE co.ID = ? OR co.ID = ?
        """, (order_id, f'`{order_id}'))

        row = cur.fetchone()
        if not row:
            conn.close()
            logger.warning(f"Order not found: {order_id}")
            return jsonify({'error': 'Order not found'}), 404

        status_code = (row[15] or '').strip()
        status_info = STATUS_MAP.get(status_code, {'label': status_code or 'Unknown', 'color': 'status-unknown'})
        currency = row[18] or 'SAR'
        total_ordered = float(row[16] or 0)
        total_shipped = float(row[17] or 0)
        shipped_pct = round((total_shipped / total_ordered * 100) if total_ordered > 0 else 0, 1)

        header = {
            'id': (row[0] or '').strip().lstrip('`'),
            'customer_id': row[1],
            'customer_name': row[37],
            'customer_po_ref': row[2],
            'contact': {
                'name': f"{row[3] or ''} {row[4] or ''}".strip(),
                'position': row[8],
                'phone': row[5],
                'mobile': row[6],
                'email': row[7],
            },
            'dates': {
                'order': fmt_date(row[9]),
                'desired_ship': fmt_date(row[10]),
                'promise': fmt_date(row[11]),
                'last_shipped': fmt_date(row[12]),
                'created': fmt_date(row[13]),
                'status_eff': fmt_date(row[14]),
            },
            'status_code': status_code,
            'status_label': status_info['label'],
            'status_color': status_info['color'],
            'financials': {
                'total_ordered': fmt_currency(total_ordered, currency),
                'total_shipped': fmt_currency(total_shipped, currency),
                'total_ordered_raw': total_ordered,
                'total_shipped_raw': total_shipped,
                'shipped_pct': shipped_pct,
                'currency': currency,
                'sell_rate': row[19],
                'buy_rate': row[20],
            },
            'terms': {
                'description': row[21],
                'net_days': row[22],
                'disc_percent': row[23],
            },
            'sales': {
                'rep': row[24],
                'territory': row[25],
            },
            'shipping': {
                'ship_via': row[26],
                'fob': row[27],
                'freight_terms': row[28],
                'backorder': row[29],
                'warehouse': row[31],
                'site': row[32],
            },
            'address': {
                'addr_1': row[38],
                'addr_2': row[39],
                'city': row[40],
                'country': row[41],
            },
            'user_fields': {
                'user_1': row[33],
                'user_2': row[34],
                'user_3': row[35],
            },
            'entered_by': row[36],
            'type': row[42],
            'order_type': row[43],
        }

        # --- Line Items ---
        where_clauses = ["CUST_ORDER_ID = ? OR CUST_ORDER_ID = ?"]
        params = [order_id, f'`{order_id}']

        if part_id:
            where_clauses.append("PART_ID LIKE ?")
            params.append(f'%{part_id}%')

        if line_status_filter:
            placeholders = ','.join('?' for _ in line_status_filter)
            where_clauses.append(f"LINE_STATUS IN ({placeholders})")
            params.extend(line_status_filter)

        where_sql = " AND ".join(where_clauses)

        cur.execute(f"""
            SELECT
                LINE_NO, PART_ID, CUSTOMER_PART_ID,
                LINE_STATUS, ORDER_QTY, TOTAL_SHIPPED_QTY,
                UNIT_PRICE, TOTAL_AMT_ORDERED, TOTAL_AMT_SHIPPED,
                DESIRED_SHIP_DATE, LAST_SHIPPED_DATE, PROMISE_DATE,
                PRODUCT_CODE, MISC_REFERENCE,
                TRADE_DISC_PERCENT, COMMISSION_PCT,
                WAREHOUSE_ID, TYPE
            FROM CUST_ORDER_LINE
            WHERE {where_sql}
            ORDER BY LINE_NO
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
        """, params + [offset, limit])

        line_rows = cur.fetchall()
        conn.close()

        lines = []
        for lr in line_rows:
            ls_code = (lr[3] or '').strip()
            ls_info = LINE_STATUS_MAP.get(ls_code, {'label': ls_code or '—', 'color': 'status-unknown'})
            lines.append({
                'line_no': lr[0],
                'part_id': lr[1],
                'customer_part_id': lr[2],
                'status_code': ls_code,
                'status_label': ls_info['label'],
                'status_color': ls_info['color'],
                'order_qty': float(lr[4] or 0),
                'shipped_qty': float(lr[5] or 0),
                'unit_price': fmt_currency(lr[6], currency),
                'total_ordered': fmt_currency(lr[7], currency),
                'total_shipped': fmt_currency(lr[8], currency),
                'desired_ship_date': fmt_date(lr[9]),
                'last_shipped_date': fmt_date(lr[10]),
                'promise_date': fmt_date(lr[11]),
                'product_code': lr[12],
                'description': lr[13],
                'disc_pct': lr[14],
                'commission_pct': lr[15],
                'warehouse': lr[16],
                'type': lr[17],
            })

        # Check if order is complete (100% fulfilled)
        if is_order_complete(lines):
            # Override status to Complete if all lines are fully shipped
            header['status_code'] = 'COMPLETE'
            header['status_label'] = STATUS_MAP['COMPLETE']['label']
            header['status_color'] = STATUS_MAP['COMPLETE']['color']

        logger.info(f"Order detail loaded: {order_id}")
        return jsonify({'header': header, 'lines': lines})
    
    except pyodbc.DatabaseError as e:
        logger.error(f"Database error in order_detail: {str(e)}")
        return jsonify({'error': 'Database connection failed'}), 503
    except Exception as e:
        logger.error(f"Error in order_detail: {str(e)}")
        return jsonify({'error': 'An unexpected error occurred'}), 500

@app.route('/api/orders/search')
def orders_search():
    try:
        q = request.args.get('q', '').strip()
        if len(q) < 2:
            return jsonify([])
        
        # Get filter parameters
        status_filter = request.args.get('status', '').split(',') if request.args.get('status') else []
        order_date_from = request.args.get('order_date_from')
        order_date_to = request.args.get('order_date_to')
        salesrep = request.args.get('salesrep')
        territory = request.args.get('territory')
        currency = request.args.get('currency')
        total_ordered_min = request.args.get('total_ordered_min')
        total_ordered_max = request.args.get('total_ordered_max')
        order_type = request.args.get('order_type')
        backorder = request.args.get('backorder')
        limit = int(request.args.get('limit', 50))
        offset = int(request.args.get('offset', 0))

        conn = get_conn()
        cur = conn.cursor()

        # Build WHERE clause dynamically
        where_clauses = ["(co.CUSTOMER_PO_REF LIKE ? OR c.NAME LIKE ?)"]
        params = [f'%{q}%', f'%{q}%']

        if status_filter:
            placeholders = ','.join('?' for _ in status_filter)
            where_clauses.append(f"co.STATUS IN ({placeholders})")
            params.extend(status_filter)

        if order_date_from and order_date_to:
            where_clauses.append("co.ORDER_DATE BETWEEN ? AND ?")
            params.extend([order_date_from, order_date_to])
        elif order_date_from:
            where_clauses.append("co.ORDER_DATE >= ?")
            params.append(order_date_from)
        elif order_date_to:
            where_clauses.append("co.ORDER_DATE <= ?")
            params.append(order_date_to)

        if salesrep:
            where_clauses.append("co.SALESREP_ID = ?")
            params.append(salesrep)

        if territory:
            where_clauses.append("co.TERRITORY LIKE ?")
            params.append(f'%{territory}%')

        if currency:
            where_clauses.append("co.CURRENCY_ID = ?")
            params.append(currency)

        if total_ordered_min and total_ordered_max:
            where_clauses.append("co.TOTAL_AMT_ORDERED BETWEEN ? AND ?")
            params.extend([float(total_ordered_min), float(total_ordered_max)])
        elif total_ordered_min:
            where_clauses.append("co.TOTAL_AMT_ORDERED >= ?")
            params.append(float(total_ordered_min))
        elif total_ordered_max:
            where_clauses.append("co.TOTAL_AMT_ORDERED <= ?")
            params.append(float(total_ordered_max))

        if order_type:
            where_clauses.append("co.ORDER_TYPE = ?")
            params.append(order_type)

        if backorder:
            where_clauses.append("co.BACKORDER_FLAG = ?")
            params.append(backorder)

        where_sql = " AND ".join(where_clauses)

        cur.execute(f"""
            SELECT TOP 100
                co.ID,
                co.CUSTOMER_ID,
                co.CUSTOMER_PO_REF,
                co.ORDER_DATE,
                co.DESIRED_SHIP_DATE,
                co.PROMISE_DATE,
                co.LAST_SHIPPED_DATE,
                co.STATUS,
                co.TOTAL_AMT_ORDERED,
                co.TOTAL_AMT_SHIPPED,
                co.CURRENCY_ID,
                co.SALESREP_ID,
                co.TERRITORY,
                co.USER_1,
                c.NAME AS CUSTOMER_NAME
            FROM CUSTOMER_ORDER co
            LEFT JOIN CUSTOMER c ON co.CUSTOMER_ID = c.ID
            WHERE {where_sql}
            ORDER BY co.ORDER_DATE DESC
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
        """, params + [offset, limit])
        rows = cur.fetchall()
        conn.close()
        
        orders = []
        for r in rows:
            status_code = (r[7] or '').strip()
            status_info = STATUS_MAP.get(status_code, {'label': status_code or 'Unknown', 'color': 'status-unknown'})
            currency_val = r[10] or 'SAR'
            orders.append({
                'id': (r[0] or '').strip().lstrip('`'),
                'customer_id': r[1],
                'customer_name': r[15],
                'customer_po_ref': r[2],
                'order_date': fmt_date(r[3]),
                'desired_ship_date': fmt_date(r[4]),
                'promise_date': fmt_date(r[5]),
                'last_shipped_date': fmt_date(r[6]),
                'status_code': status_code,
                'status_label': status_info['label'],
                'status_color': status_info['color'],
                'total_ordered': fmt_currency(r[8], currency_val),
                'total_shipped': fmt_currency(r[9], currency_val),
                'total_ordered_raw': float(r[8] or 0),
                'total_shipped_raw': float(r[9] or 0),
                'currency': currency_val,
                'salesrep': r[11],
                'territory': r[12],
                'site': r[13],
            })
        return jsonify(orders)
    except pyodbc.DatabaseError as e:
        logger.error(f"Database error in orders_search: {str(e)}")
        return jsonify({'error': 'Database connection failed'}), 503
    except Exception as e:
        logger.error(f"Error in orders_search: {str(e)}")
        return jsonify({'error': 'An unexpected error occurred'}), 500

@app.route('/api/salesreps')
def salesreps():
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT DISTINCT SALESREP_ID, TERRITORY
            FROM CUSTOMER_ORDER
            WHERE SALESREP_ID IS NOT NULL OR TERRITORY IS NOT NULL
            ORDER BY SALESREP_ID, TERRITORY
        """)
        rows = cur.fetchall()
        conn.close()
        
        reps = []
        territories = []
        for r in rows:
            if r[0] and r[0] not in [rep['id'] for rep in reps]:
                reps.append({'id': r[0], 'label': r[0]})
            if r[1] and r[1] not in territories:
                territories.append(r[1])
        
        return jsonify({'salesreps': reps, 'territories': territories})
    except pyodbc.DatabaseError as e:
        logger.error(f"Database error in salesreps: {str(e)}")
        return jsonify({'error': 'Database connection failed'}), 503
    except Exception as e:
        logger.error(f"Error in salesreps: {str(e)}")
        return jsonify({'error': 'An unexpected error occurred'}), 500

@app.route('/api/outstanding-payments')
def outstanding_payments():
    try:
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        limit = int(request.args.get('limit', 50))
        offset = int(request.args.get('offset', 0))

        conn = get_conn()
        cur = conn.cursor()

        where_clauses = ["t.AMT_OUTSTANDING > 0"]
        params = []

        if date_from and date_to:
            where_clauses.append("co.ORDER_DATE BETWEEN ? AND ?")
            params.extend([date_from, date_to])
        elif date_from:
            where_clauses.append("co.ORDER_DATE >= ?")
            params.append(date_from)
        elif date_to:
            where_clauses.append("co.ORDER_DATE <= ?")
            params.append(date_to)

        where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"

        cur.execute(f"""
            SELECT
                t.ID,
                t.CUSTOMER_PO_REF,
                t.CONTACT_FIRST_NAME,
                t.SHIP_VIA,
                t.SALESREP_ID,
                t.FREIGHT_TERMS,
                t.TERMS_DESCRIPTION,
                t.STATUS,
                t.TOTAL_AMT_ORDERED,
                t.TOTAL_AMT_SHIPPED,
                t.AMT_OUTSTANDING,
                co.ORDER_DATE,
                co.CURRENCY_ID,
                c.NAME AS CUSTOMER_NAME
            FROM (
                SELECT 
                    ID,
                    CUSTOMER_PO_REF,
                    CONTACT_FIRST_NAME,
                    SHIP_VIA,
                    SALESREP_ID,
                    FREIGHT_TERMS,
                    TERMS_DESCRIPTION,
                    STATUS,
                    TOTAL_AMT_ORDERED,
                    TOTAL_AMT_SHIPPED,
                    (ISNULL(TOTAL_AMT_ORDERED, 0) - ISNULL(TOTAL_AMT_SHIPPED, 0)) AS AMT_OUTSTANDING
                FROM CUSTOMER_ORDER
            ) t
            LEFT JOIN CUSTOMER_ORDER co ON t.ID = co.ID
            LEFT JOIN CUSTOMER c ON co.CUSTOMER_ID = c.ID
            WHERE {where_sql}
            ORDER BY co.ORDER_DATE DESC
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
        """, params + [offset, limit])
        rows = cur.fetchall()
        conn.close()
        
        payments = []
        for r in rows:
            status_code = (r[7] or '').strip()
            status_info = STATUS_MAP.get(status_code, {'label': status_code or 'Unknown', 'color': 'status-unknown'})
            currency = r[12] or 'SAR'
            payments.append({
                'id': (r[0] or '').strip().lstrip('`'),
                'customer_po_ref': r[1],
                'contact_first_name': r[2],
                'ship_via': r[3],
                'salesrep': r[4],
                'freight_terms': r[5],
                'terms_description': r[6],
                'status_code': status_code,
                'status_label': status_info['label'],
                'status_color': status_info['color'],
                'total_ordered': fmt_currency(r[8], currency),
                'total_shipped': fmt_currency(r[9], currency),
                'amt_outstanding': fmt_currency(r[10], currency),
                'amt_outstanding_raw': float(r[10] or 0),
                'order_date': fmt_date(r[11]),
                'currency': currency,
                'customer_name': r[13],
            })
        return jsonify(payments)
    except ValueError as e:
        logger.error(f"Invalid parameter: {str(e)}")
        return jsonify({'error': 'Invalid parameter format'}), 400
    except pyodbc.DatabaseError as e:
        logger.error(f"Database error in outstanding_payments: {str(e)}")
        return jsonify({'error': 'Database connection failed'}), 503
    except Exception as e:
        logger.error(f"Error in outstanding_payments: {str(e)}")
        return jsonify({'error': 'An unexpected error occurred'}), 500

# --- GL ROUTES ---

@app.route('/gl')
def gl_index():
    """GL Lookup page route."""
    return render_template('indexgl.html')

@app.route('/api/gl/search')
def gl_search():
    """GL search endpoint."""
    try:
        q            = request.args.get('q', '').strip()
        start_year   = int(request.args.get('start_year',  2025))
        start_period = int(request.args.get('start_period', 1))
        end_year     = int(request.args.get('end_year',    2026))
        end_period   = int(request.args.get('end_period',  12))

        # Input validation
        if not q:
            return jsonify({'error': 'Please enter an Account ID or Description'}), 400
        
        if len(q) < 2:
            return jsonify({'error': 'Search query must be at least 2 characters'}), 400
        
        if start_year < 2000 or start_year > 2099 or end_year < 2000 or end_year > 2099:
            return jsonify({'error': 'Invalid year range. Please use years between 2000-2099'}), 400
        
        if start_period < 1 or start_period > 12 or end_period < 1 or end_period > 12:
            return jsonify({'error': 'Invalid period. Please use periods between 1-12'}), 400
        
        # Validate date range logic
        if (start_year > end_year) or (start_year == end_year and start_period > end_period):
            return jsonify({'error': 'Start date must be before or equal to end date'}), 400

        logger.info(f"Searching GL data: query='{q}', range={start_year}/P{start_period}-{end_year}/P{end_period}")
        
        sql, params = build_gl_query(q, start_year, start_period, end_year, end_period)
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(sql, params)
        rows = cur.fetchall()
        conn.close()
        
        # Convert to list of dictionaries with proper column names
        columns = [column[0] for column in cur.description]
        data = [dict(zip(columns, row)) for row in rows]
        
        logger.info(f"Query returned {len(data)} records")
        return jsonify({'data': data, 'count': len(data)})
    except ValueError as e:
        logger.error(f"Invalid parameter: {str(e)}")
        return jsonify({'error': 'Invalid parameter format'}), 400
    except pyodbc.DatabaseError as e:
        logger.error(f"Database error in gl_search: {str(e)}")
        return jsonify({'error': 'Database connection failed'}), 503
    except Exception as e:
        logger.error(f"Error in gl_search: {str(e)}")
        return jsonify({'error': 'An unexpected error occurred'}), 500

@app.route('/api/gl/export')
def gl_export():
    """GL export endpoint."""
    try:
        q            = request.args.get('q', '').strip()
        start_year   = int(request.args.get('start_year',  2025))
        start_period = int(request.args.get('start_period', 1))
        end_year     = int(request.args.get('end_year',    2026))
        end_period   = int(request.args.get('end_period',  12))

        # Input validation (same as search)
        if not q:
            return jsonify({'error': 'No search query'}), 400
        
        if len(q) < 2:
            return jsonify({'error': 'Search query must be at least 2 characters'}), 400
        
        if start_year < 2000 or start_year > 2099 or end_year < 2000 or end_year > 2099:
            return jsonify({'error': 'Invalid year range. Please use years between 2000-2099'}), 400
        
        if start_period < 1 or start_period > 12 or end_period < 1 or end_period > 12:
            return jsonify({'error': 'Invalid period. Please use periods between 1-12'}), 400
        
        if (start_year > end_year) or (start_year == end_year and start_period > end_period):
            return jsonify({'error': 'Start date must be before or equal to end date'}), 400

        logger.info(f"Exporting GL data: query='{q}', range={start_year}/P{start_period}-{end_year}/P{end_period}")
        
        sql, params = build_gl_query(q, start_year, start_period, end_year, end_period)
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(sql, params)
        rows = cur.fetchall()
        conn.close()
        
        if not rows:
            return jsonify({'error': 'No data found for the specified criteria'}), 404

        # Convert to list of dictionaries with proper column names
        columns = [column[0] for column in cur.description]
        data = [dict(zip(columns, row)) for row in rows]
        
        # Filter out internal columns
        filtered_data = []
        for row in data:
            filtered_row = {k: v for k, v in row.items() if k not in ['Fiscal_Year', 'Fiscal_Period']}
            filtered_data.append(filtered_row)

        wb = Workbook()
        ws = wb.active
        ws.title = 'General Ledger'

        # Palette: Obsidian Gold
        DARK    = '0C0C0F'
        SURFACE = '1C1917'
        GOLD    = 'BA7517'
        LIGHT   = 'FAEEDA'
        WHITE   = 'FFFFFF'

        thin = Side(style='thin', color='333333')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Title row
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = f'General Ledger \u2014 Account: {q}'
        title_cell.font = Font(name='Calibri', bold=True, size=14, color=LIGHT)
        title_cell.fill = PatternFill('solid', start_color=DARK)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 32

        # Subtitle
        ws.merge_cells('A2:G2')
        sub = ws['A2']
        sub.value = f'Fiscal Period {start_year}/P{start_period} \u2192 {end_year}/P{end_period}'
        sub.font = Font(name='Calibri', size=10, color=GOLD)
        sub.fill = PatternFill('solid', start_color=SURFACE)
        sub.alignment = Alignment(horizontal='center')
        ws.row_dimensions[2].height = 20

        # Headers
        headers = ['Account ID', 'Description', 'Period', 'Balance', 'Budget', 'Cumulative Budget', 'Variance']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = Font(name='Calibri', bold=True, size=10, color=LIGHT)
            cell.fill = PatternFill('solid', start_color=GOLD)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        ws.row_dimensions[3].height = 20

        # Data rows
        money_fmt = '#,##0.00;(#,##0.00);"-'
        for r_idx, row_data in enumerate(filtered_data, 4):
            bg = 'FFFFFF' if r_idx % 2 == 0 else 'F7F0E6'
            # Order columns to match headers
            row_values = [
                row_data.get('Account_ID', ''),
                row_data.get('Description', ''),
                row_data.get('Period', ''),
                row_data.get('Balance', 0),
                row_data.get('Budget', 0),
                row_data.get('Cumulative_Budget', 0),
                row_data.get('Variance', 0)
            ]
            for c_idx, val in enumerate(row_values, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.fill = PatternFill('solid', start_color=bg)
                cell.font = Font(name='Calibri', size=10)
                cell.border = border
                if c_idx in (4, 5, 6, 7):  # numeric cols
                    cell.number_format = money_fmt
                    cell.alignment = Alignment(horizontal='right')
                elif c_idx == 1:
                    cell.font = Font(name='Calibri', size=10, bold=True, color=GOLD[2:] if len(GOLD)>6 else GOLD)
                    cell.alignment = Alignment(horizontal='left')

        # Column widths
        col_widths = [14, 36, 12, 18, 16, 20, 18]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Totals row
        last_data = 3 + len(filtered_data)
        total_row = last_data + 1
        ws.merge_cells(f'A{total_row}:C{total_row}')
        t = ws.cell(row=total_row, column=1, value='TOTAL')
        t.font = Font(name='Calibri', bold=True, color=LIGHT)
        t.fill = PatternFill('solid', start_color=DARK)
        t.alignment = Alignment(horizontal='right')

        for col in (4, 5, 7):
            col_letter = get_column_letter(col)
            cell = ws.cell(row=total_row, column=col,
                           value=f'=SUM({col_letter}4:{col_letter}{last_data})')
            cell.font = Font(name='Calibri', bold=True, color=LIGHT)
            cell.fill = PatternFill('solid', start_color=DARK)
            cell.number_format = money_fmt
            cell.alignment = Alignment(horizontal='right')
            cell.border = border

        ws.cell(row=total_row, column=6).fill = PatternFill('solid', start_color=DARK)

        ws.freeze_panes = 'A4'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f'GL_{q.replace(" ","_")}_{start_year}P{start_period}-{end_year}P{end_period}.xlsx'
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except ValueError as e:
        logger.error(f"Invalid parameter: {str(e)}")
        return jsonify({'error': 'Invalid parameter format'}), 400
    except pyodbc.DatabaseError as e:
        logger.error(f"Database error in gl_export: {str(e)}")
        return jsonify({'error': 'Database connection failed'}), 503
    except Exception as e:
        logger.error(f"Error in gl_export: {str(e)}")
        return jsonify({'error': 'An unexpected error occurred'}), 500

@app.route('/suppliers/dashboard')
def suppliers_dashboard():
    return render_template('suppliers_dashboard.html')

@app.route('/api/suppliers/summary')
def supplier_summary():
    try:
        conn = get_conn()
        cur = conn.cursor()

        cur.execute("""
            SELECT
                v.ID AS vendor_id,
                v.NAME AS vendor_name,
                COUNT(po.ID) AS number_of_pos,
                SUM(po.TOTAL_AMT_ORDERED) AS total_spend,
                SUM(po.TOTAL_AMT_RECVD) AS total_received,
                MAX(po.ORDER_DATE) AS last_order_date
            FROM VENDOR v
            JOIN PURCHASE_ORDER po
                ON po.VENDOR_ID = v.ID
            WHERE po.ORDER_DATE >= DATEADD(MONTH, -6, GETDATE())
            GROUP BY v.ID, v.NAME
            ORDER BY total_spend DESC
        """)

        rows = cur.fetchall()
        conn.close()

        result = []
        for r in rows:
            result.append({
                "vendor_id": r[0],
                "vendor_name": r[1],
                "number_of_pos": r[2],
                "total_spend": float(r[3] or 0),
                "total_received": float(r[4] or 0),
                "last_order_date": fmt_date(r[5])
            })

        return jsonify(result)

    except Exception as e:
        logger.error(f"Supplier summary error: {str(e)}")
        return jsonify({"error": "Failed to load supplier summary"}), 500

@app.route('/api/suppliers/top10')
def suppliers_top10():
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT TOP 10
            v.NAME,
            SUM(po.TOTAL_AMT_ORDERED) AS SPEND
        FROM VENDOR v
        JOIN PURCHASE_ORDER po ON po.VENDOR_ID = v.ID
        GROUP BY v.NAME
        ORDER BY SPEND DESC
    """)

    rows = cur.fetchall()

    return jsonify([
        {"name": r[0], "spend": float(r[1] or 0)}
        for r in rows
    ])
    
@app.route('/api/suppliers/monthly-po')
def monthly_po():
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            FORMAT(ORDER_DATE, 'yyyy-MM') AS month,
            COUNT(*) AS po_count
        FROM PURCHASE_ORDER
        WHERE ORDER_DATE >= DATEADD(MONTH, -6, GETDATE())
        GROUP BY FORMAT(ORDER_DATE, 'yyyy-MM')
        ORDER BY month
    """)

    rows = cur.fetchall()
    conn.close()

    return jsonify([
        {"month": r[0], "po_count": r[1]}
        for r in rows
    ])

@app.route('/api/suppliers/monthly')
def monthly_pos():
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            FORMAT(ORDER_DATE, 'yyyy-MM') AS month,
            COUNT(*)
        FROM PURCHASE_ORDER
        WHERE ORDER_DATE >= DATEADD(MONTH, -6, GETDATE())
        GROUP BY FORMAT(ORDER_DATE, 'yyyy-MM')
        ORDER BY month
    """)

    rows = cur.fetchall()

    return jsonify([
        {"month": r[0], "count": r[1]}
        for r in rows
    ])

@app.route('/api/suppliers/export')
def supplier_export():
    try:
        conn = get_conn()
        cur = conn.cursor()

        cur.execute("""
            SELECT
                v.ID,
                v.NAME,
                COUNT(po.ID) AS POS,
                SUM(po.TOTAL_AMT_ORDERED) AS SPEND,
                SUM(po.TOTAL_AMT_RECVD) AS RECEIVED,
                MAX(po.ORDER_DATE) AS LAST_ORDER
            FROM VENDOR v
            JOIN PURCHASE_ORDER po
                ON po.VENDOR_ID = v.ID
            WHERE po.ORDER_DATE >= DATEADD(MONTH, -6, GETDATE())
            GROUP BY v.ID, v.NAME
            ORDER BY SPEND DESC
        """)

        rows = cur.fetchall()
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Suppliers"

        headers = ["Vendor ID", "Vendor Name", "PO Count", "Total Spend", "Total Received", "Last Order"]

        for c, h in enumerate(headers, 1):
            ws.cell(row=1, column=c, value=h).font = Font(bold=True)

        for i, r in enumerate(rows, 2):
            ws.cell(i, 1, r[0])
            ws.cell(i, 2, r[1])
            ws.cell(i, 3, r[2])
            ws.cell(i, 4, float(r[3] or 0))
            ws.cell(i, 5, float(r[4] or 0))
            ws.cell(i, 6, fmt_date(r[5]))

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return send_file(
            buf,
            as_attachment=True,
            download_name="supplier_report.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"Supplier export error: {str(e)}")
        return jsonify({"error": "Export failed"}), 500

# Add favicon route to prevent 404 errors
@app.route('/favicon.ico')
def favicon():
    return '', 204  # Return No Content


if __name__ == '__main__':
    logger.info(f"Starting Flask app on {FLASK_HOST}:{FLASK_PORT}")
    app.run(host=FLASK_HOST, port=FLASK_PORT, debug=FLASK_DEBUG)