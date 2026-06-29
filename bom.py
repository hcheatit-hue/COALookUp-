import os
import logging
from decimal import Decimal

import pyodbc
from dotenv import load_dotenv
from flask import Flask, request, render_template_string, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

load_dotenv()

app = Flask(__name__)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


# =========================
# Database connection
# =========================

def get_connection():
    """
    Reads SQL Server connection settings from .env.
    Example .env:

    DB_SERVER=HCHDB01\\SSINFOR
    DB_NAME=SSHCHEAT
    DB_USER=your_username
    DB_PASSWORD=your_password
    DB_DRIVER={ODBC Driver 17 for SQL Server}
    DB_TRUST_CERT=yes
    DB_ENCRYPT=no
    """
    server = os.getenv("DB_SERVER")
    database = os.getenv("DB_NAME")
    username = os.getenv("DB_USER")
    password = os.getenv("DB_PASSWORD")
    driver = os.getenv("DB_DRIVER", "{ODBC Driver 17 for SQL Server}")
    trust_cert = os.getenv("DB_TRUST_CERT", "yes")
    encrypt = os.getenv("DB_ENCRYPT", "no")

    if not all([server, database]):
        raise RuntimeError("Missing DB_SERVER or DB_NAME in .env")

    # SQL Authentication
    if username and password:
        connection_string = (
            f"DRIVER={driver};"
            f"SERVER={server};"
            f"DATABASE={database};"
            f"UID={username};"
            f"PWD={password};"
            f"TrustServerCertificate={trust_cert};"
            f"Encrypt={encrypt};"
        )

    # Windows Integrated Authentication
    else:
        connection_string = (
            f"DRIVER={driver};"
            f"SERVER={server};"
            f"DATABASE={database};"
            f"Trusted_Connection=yes;"
            f"TrustServerCertificate={trust_cert};"
            f"Encrypt={encrypt};"
        )

    return pyodbc.connect(connection_string)


# =========================
# Helpers
# =========================

def clean_decimal(value):
    if value is None:
        return 0
    if isinstance(value, Decimal):
        return float(value)
    return value


def rows_to_dicts(cursor):
    columns = [column[0] for column in cursor.description]
    records = []

    for row in cursor.fetchall():
        item = {}
        for index, column in enumerate(columns):
            item[column] = clean_decimal(row[index])
        records.append(item)

    return records


# =========================
# SQL queries
# =========================

ALL_REQUIREMENTS_SQL = """
SELECT
    r.WORKORDER_BASE_ID AS WorkOrder,
    r.PART_ID AS PartID,
    p.DESCRIPTION AS Description,
    r.STATUS AS Status,
    r.CALC_QTY AS RequiredQty,
    ISNULL(r.ISSUED_QTY, 0) AS IssuedQty,
    r.CALC_QTY - ISNULL(r.ISSUED_QTY, 0) AS NotIssuedQty,
    CASE
        WHEN ISNULL(r.ISSUED_QTY, 0) = 0
            THEN 'NOT ISSUED'
        WHEN ISNULL(r.ISSUED_QTY, 0) < r.CALC_QTY
            THEN 'PARTIALLY ISSUED'
        WHEN ISNULL(r.ISSUED_QTY, 0) >= r.CALC_QTY
            THEN 'FULLY ISSUED'
    END AS IssueStatus
FROM REQUIREMENT r
LEFT JOIN PART p
    ON p.ID = r.PART_ID
WHERE r.WORKORDER_BASE_ID = ?
ORDER BY
    CASE
        WHEN ISNULL(r.ISSUED_QTY, 0) = 0 THEN 1
        WHEN ISNULL(r.ISSUED_QTY, 0) < r.CALC_QTY THEN 2
        WHEN ISNULL(r.ISSUED_QTY, 0) >= r.CALC_QTY THEN 3
        ELSE 4
    END,
    r.PART_ID;
"""

SHORT_ISSUED_SQL = """
SELECT
    r.WORKORDER_BASE_ID AS WorkOrder,
    r.PART_ID AS PartID,
    p.DESCRIPTION AS Description,
    r.STATUS AS Status,
    r.CALC_QTY AS RequiredQty,
    ISNULL(r.ISSUED_QTY, 0) AS IssuedQty,
    r.CALC_QTY - ISNULL(r.ISSUED_QTY, 0) AS NotIssuedQty,
    CASE
        WHEN ISNULL(r.ISSUED_QTY, 0) = 0
            THEN 'NOT ISSUED'
        WHEN ISNULL(r.ISSUED_QTY, 0) < r.CALC_QTY
            THEN 'PARTIALLY ISSUED'
        WHEN ISNULL(r.ISSUED_QTY, 0) >= r.CALC_QTY
            THEN 'FULLY ISSUED'
    END AS IssueStatus
FROM REQUIREMENT r
LEFT JOIN PART p
    ON p.ID = r.PART_ID
WHERE r.WORKORDER_BASE_ID = ?
  AND ISNULL(r.ISSUED_QTY, 0) < r.CALC_QTY
ORDER BY r.PART_ID;
"""


def get_workorder_requirements(workorder, status_filter="ALL"):
    """
    status_filter options:
    ALL, NOT ISSUED, PARTIALLY ISSUED, FULLY ISSUED, SHORT
    SHORT means NOT ISSUED + PARTIALLY ISSUED.
    """
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(ALL_REQUIREMENTS_SQL, workorder)
        rows = rows_to_dicts(cursor)

    if status_filter == "SHORT":
        return [row for row in rows if row.get("IssueStatus") in ["NOT ISSUED", "PARTIALLY ISSUED"]]

    if status_filter in ["NOT ISSUED", "PARTIALLY ISSUED", "FULLY ISSUED"]:
        return [row for row in rows if row.get("IssueStatus") == status_filter]

    return rows


def calculate_summary(rows):
    total_required = sum(float(row.get("RequiredQty") or 0) for row in rows)
    total_issued = sum(float(row.get("IssuedQty") or 0) for row in rows)
    total_outstanding = sum(float(row.get("NotIssuedQty") or 0) for row in rows)

    issued_percent = 0
    if total_required > 0:
        issued_percent = round((total_issued / total_required) * 100, 2)

    return {
        "total": len(rows),
        "not_issued": sum(1 for row in rows if row.get("IssueStatus") == "NOT ISSUED"),
        "partial": sum(1 for row in rows if row.get("IssueStatus") == "PARTIALLY ISSUED"),
        "full": sum(1 for row in rows if row.get("IssueStatus") == "FULLY ISSUED"),
        "short": sum(1 for row in rows if row.get("IssueStatus") in ["NOT ISSUED", "PARTIALLY ISSUED"]),
        "total_required": total_required,
        "total_issued": total_issued,
        "total_outstanding": total_outstanding,
        "issued_percent": issued_percent,
    }


def create_excel_file(workorder, rows, summary):
    wb = Workbook()
    ws = wb.active
    ws.title = "Issue Status"

    ws["A1"] = "Work Order Issue Status"
    ws["A1"].font = Font(size=16, bold=True)
    ws.merge_cells("A1:H1")

    ws["A3"] = "Work Order"
    ws["B3"] = workorder
    ws["A4"] = "Total Lines"
    ws["B4"] = summary["total"]
    ws["A5"] = "Not Issued / Partial"
    ws["B5"] = summary["short"]
    ws["A6"] = "Fully Issued"
    ws["B6"] = summary["full"]
    ws["D3"] = "Total Required Qty"
    ws["E3"] = summary["total_required"]
    ws["D4"] = "Total Issued Qty"
    ws["E4"] = summary["total_issued"]
    ws["D5"] = "Total Outstanding Qty"
    ws["E5"] = summary["total_outstanding"]
    ws["D6"] = "% Issued"
    ws["E6"] = summary["issued_percent"]

    headers = [
        "Work Order", "Part ID", "Description", "Status",
        "Required Qty", "Issued Qty", "Not Issued Qty", "Issue Status"
    ]
    start_row = 8

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9DEE5"),
        right=Side(style="thin", color="D9DEE5"),
        top=Side(style="thin", color="D9DEE5"),
        bottom=Side(style="thin", color="D9DEE5"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_index, row in enumerate(rows, start_row + 1):
        values = [
            row.get("WorkOrder"),
            row.get("PartID"),
            row.get("Description"),
            row.get("Status"),
            row.get("RequiredQty"),
            row.get("IssuedQty"),
            row.get("NotIssuedQty"),
            row.get("IssueStatus"),
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_index, column=col, value=value)
            cell.border = thin_border

        issue_status = row.get("IssueStatus")
        if issue_status == "NOT ISSUED":
            fill = PatternFill("solid", fgColor="FEE4E2")
        elif issue_status == "PARTIALLY ISSUED":
            fill = PatternFill("solid", fgColor="FEF0C7")
        elif issue_status == "FULLY ISSUED":
            fill = PatternFill("solid", fgColor="DCFAE6")
        else:
            fill = None

        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_index, column=col).fill = fill

    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 45)

    ws.freeze_panes = "A9"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# =========================
# Frontend template
# =========================

HTML = """
<!doctype html>
<html lang="en">
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Work Order Issue Status</title>
    <style>
        :root {
            --bg: #f4f6f8;
            --card: #ffffff;
            --text: #17202a;
            --muted: #6b7280;
            --border: #d9dee5;
            --primary: #1f4e79;
            --primary-dark: #173a5a;
            --danger: #b42318;
            --warning: #b54708;
            --success: #067647;
            --shadow: 0 10px 25px rgba(15, 23, 42, 0.08);
        }

        * {
            box-sizing: border-box;
        }

        body {
            margin: 0;
            font-family: Arial, Helvetica, sans-serif;
            background: var(--bg);
            color: var(--text);
        }

        .page {
            max-width: 1250px;
            margin: 0 auto;
            padding: 28px;
        }

        .header {
            display: flex;
            justify-content: space-between;
            gap: 16px;
            align-items: center;
            margin-bottom: 22px;
        }

        .header h1 {
            margin: 0;
            font-size: 28px;
        }

        .header p {
            margin: 6px 0 0;
            color: var(--muted);
        }

        .card {
            background: var(--card);
            border: 1px solid var(--border);
            border-radius: 16px;
            box-shadow: var(--shadow);
            padding: 18px;
            margin-bottom: 18px;
        }

        .search-form {
            display: grid;
            grid-template-columns: 1fr auto auto;
            gap: 12px;
            align-items: center;
        }

        input[type="text"] {
            width: 100%;
            padding: 13px 14px;
            border-radius: 10px;
            border: 1px solid var(--border);
            font-size: 16px;
        }

        .checkbox-wrap {
            display: flex;
            align-items: center;
            gap: 8px;
            color: var(--text);
            white-space: nowrap;
            font-size: 14px;
        }

        button, .button-link {
            border: none;
            background: var(--primary);
            color: white;
            padding: 13px 18px;
            border-radius: 10px;
            cursor: pointer;
            font-size: 15px;
            text-decoration: none;
            display: inline-block;
        }

        button:hover, .button-link:hover {
            background: var(--primary-dark);
        }

        .summary {
            display: grid;
            grid-template-columns: repeat(4, 1fr);
            gap: 14px;
        }

        .summary-box {
            border: 1px solid var(--border);
            border-radius: 14px;
            padding: 16px;
            background: #fbfdff;
        }

        .summary-box span {
            color: var(--muted);
            font-size: 13px;
        }

        .summary-box strong {
            display: block;
            margin-top: 6px;
            font-size: 26px;
        }

        .toolbar {
            display: flex;
            justify-content: space-between;
            gap: 12px;
            align-items: center;
            margin-bottom: 12px;
        }

        .toolbar input {
            max-width: 320px;
        }

        table {
            width: 100%;
            border-collapse: collapse;
            overflow: hidden;
            border-radius: 12px;
        }

        th, td {
            padding: 12px 10px;
            text-align: left;
            border-bottom: 1px solid var(--border);
            font-size: 14px;
        }

        th {
            background: #eef3f8;
            color: #243447;
            font-weight: 700;
        }

        tr:hover td {
            background: #fafcff;
        }

        .num {
            text-align: right;
            font-variant-numeric: tabular-nums;
        }

        .badge {
            display: inline-block;
            padding: 5px 9px;
            border-radius: 999px;
            font-size: 12px;
            font-weight: 700;
            white-space: nowrap;
        }

        .badge.not-issued {
            color: var(--danger);
            background: #fee4e2;
        }

        .badge.partial {
            color: var(--warning);
            background: #fef0c7;
        }

        .badge.full {
            color: var(--success);
            background: #dcfae6;
        }

        tr.row-not-issued td {
            background: #fff6f5;
        }

        tr.row-partial td {
            background: #fffaf0;
        }

        tr.row-full td {
            background: #f6fffa;
        }

        .actions {
            display: flex;
            gap: 10px;
            align-items: center;
            flex-wrap: wrap;
        }

        select {
            padding: 12px 14px;
            border-radius: 10px;
            border: 1px solid var(--border);
            font-size: 15px;
            background: white;
        }

        .history {
            display: flex;
            flex-wrap: wrap;
            gap: 8px;
            margin-top: 12px;
        }

        .history a {
            color: var(--primary);
            background: #eef3f8;
            text-decoration: none;
            padding: 7px 10px;
            border-radius: 999px;
            font-size: 13px;
        }

        .empty {
            color: var(--muted);
            padding: 18px;
        }

        .error {
            color: var(--danger);
            background: #fff1f0;
            border: 1px solid #fecdca;
            padding: 12px;
            border-radius: 10px;
            margin-bottom: 16px;
        }

        @media (max-width: 850px) {
            .search-form,
            .summary {
                grid-template-columns: 1fr;
            }

            .header {
                display: block;
            }

            .toolbar {
                display: block;
            }

            .toolbar input {
                max-width: 100%;
                margin-top: 10px;
            }

            table {
                display: block;
                overflow-x: auto;
                white-space: nowrap;
            }
        }
    </style>
</head>
<body>
    <div class="page">
        <div class="header">
            <div>
                <h1>Work Order Issue Status</h1>
                <p>Check BOM requirement issue status from Visual ERP.</p>
            </div>
        </div>

        <div class="card">
            <form class="search-form" method="get" action="/">
                <input
                    id="workorderInput"
                    type="text"
                    name="workorder"
                    placeholder="Enter Work Order, e.g. 254306"
                    value="{{ workorder or '' }}"
                    required
                >

                <select name="status_filter">
                    <option value="ALL" {% if status_filter == 'ALL' %}selected{% endif %}>All statuses</option>
                    <option value="SHORT" {% if status_filter == 'SHORT' %}selected{% endif %}>Not fully issued only</option>
                    <option value="NOT ISSUED" {% if status_filter == 'NOT ISSUED' %}selected{% endif %}>Not issued</option>
                    <option value="PARTIALLY ISSUED" {% if status_filter == 'PARTIALLY ISSUED' %}selected{% endif %}>Partially issued</option>
                    <option value="FULLY ISSUED" {% if status_filter == 'FULLY ISSUED' %}selected{% endif %}>Fully issued</option>
                </select>

                <div class="actions">
                    <button type="submit">Search</button>
                    {% if workorder and not error %}
                        <a class="button-link" href="/export/workorder/{{ workorder }}?status_filter={{ status_filter }}">Export Excel</a>
                    {% endif %}
                </div>
            </form>

            <div class="history" id="searchHistory"></div>
        </div>

        {% if error %}
            <div class="error">{{ error }}</div>
        {% endif %}

        {% if workorder and not error %}
            <div class="card summary">
                <div class="summary-box">
                    <span>Work Order</span>
                    <strong>{{ workorder }}</strong>
                </div>
                <div class="summary-box">
                    <span>Total Lines</span>
                    <strong>{{ summary.total }}</strong>
                </div>
                <div class="summary-box">
                    <span>Not Issued / Partial</span>
                    <strong>{{ summary.short }}</strong>
                </div>
                <div class="summary-box">
                    <span>% Issued</span>
                    <strong>{{ '%.2f'|format(summary.issued_percent or 0) }}%</strong>
                </div>
            </div>

            <div class="card summary">
                <div class="summary-box">
                    <span>Total Required Qty</span>
                    <strong>{{ '%.2f'|format(summary.total_required or 0) }}</strong>
                </div>
                <div class="summary-box">
                    <span>Total Issued Qty</span>
                    <strong>{{ '%.2f'|format(summary.total_issued or 0) }}</strong>
                </div>
                <div class="summary-box">
                    <span>Total Outstanding Qty</span>
                    <strong>{{ '%.2f'|format(summary.total_outstanding or 0) }}</strong>
                </div>
                <div class="summary-box">
                    <span>Fully Issued Lines</span>
                    <strong>{{ summary.full }}</strong>
                </div>
            </div>

            <div class="card">
                <div class="toolbar">
                    <div>
                        <strong>Results</strong>
                        <span style="color: var(--muted);">({{ rows|length }} line/s)</span>
                    </div>
                    <input id="tableSearch" type="text" placeholder="Filter by part, description, status...">
                </div>

                {% if rows %}
                    <table id="resultsTable">
                        <thead>
                            <tr>
                                <th>Work Order</th>
                                <th>Part ID</th>
                                <th>Description</th>
                                <th>Status</th>
                                <th class="num">Required Qty</th>
                                <th class="num">Issued Qty</th>
                                <th class="num">Not Issued Qty</th>
                                <th>Issue Status</th>
                            </tr>
                        </thead>
                        <tbody>
                            {% for row in rows %}
                                <tr class="{% if row.IssueStatus == 'NOT ISSUED' %}row-not-issued{% elif row.IssueStatus == 'PARTIALLY ISSUED' %}row-partial{% elif row.IssueStatus == 'FULLY ISSUED' %}row-full{% endif %}">
                                    <td>{{ row.WorkOrder }}</td>
                                    <td>{{ row.PartID }}</td>
                                    <td>{{ row.Description or '' }}</td>
                                    <td>{{ row.Status or '' }}</td>
                                    <td class="num">{{ '%.2f'|format(row.RequiredQty or 0) }}</td>
                                    <td class="num">{{ '%.2f'|format(row.IssuedQty or 0) }}</td>
                                    <td class="num">{{ '%.2f'|format(row.NotIssuedQty or 0) }}</td>
                                    <td>
                                        {% if row.IssueStatus == 'NOT ISSUED' %}
                                            <span class="badge not-issued">NOT ISSUED</span>
                                        {% elif row.IssueStatus == 'PARTIALLY ISSUED' %}
                                            <span class="badge partial">PARTIALLY ISSUED</span>
                                        {% elif row.IssueStatus == 'FULLY ISSUED' %}
                                            <span class="badge full">FULLY ISSUED</span>
                                        {% else %}
                                            {{ row.IssueStatus }}
                                        {% endif %}
                                    </td>
                                </tr>
                            {% endfor %}
                        </tbody>
                    </table>
                {% else %}
                    <div class="empty">No records found for this work order.</div>
                {% endif %}
            </div>
        {% endif %}
    </div>

    <script>
        const searchInput = document.getElementById('tableSearch');
        const table = document.getElementById('resultsTable');
        const workorderInput = document.getElementById('workorderInput');
        const historyContainer = document.getElementById('searchHistory');
        const currentWorkorder = "{{ workorder or '' }}";

        function updateSearchHistory(workorder) {
            if (!workorder) return;
            const key = 'workorderSearchHistory';
            let history = JSON.parse(localStorage.getItem(key) || '[]');
            history = history.filter(item => item !== workorder);
            history.unshift(workorder);
            history = history.slice(0, 10);
            localStorage.setItem(key, JSON.stringify(history));
        }

        function renderSearchHistory() {
            if (!historyContainer) return;
            const key = 'workorderSearchHistory';
            const history = JSON.parse(localStorage.getItem(key) || '[]');

            if (history.length === 0) {
                historyContainer.innerHTML = '';
                return;
            }

            historyContainer.innerHTML = '<strong style="font-size:13px;color:var(--muted);margin-right:4px;">Recent:</strong>' +
                history.map(item => `<a href="/?workorder=${encodeURIComponent(item)}&status_filter=ALL">${item}</a>`).join('');
        }

        updateSearchHistory(currentWorkorder);
        renderSearchHistory();

        if (searchInput && table) {
            searchInput.addEventListener('input', function () {
                const filter = this.value.toLowerCase();
                const rows = table.querySelectorAll('tbody tr');

                rows.forEach(row => {
                    const text = row.innerText.toLowerCase();
                    row.style.display = text.includes(filter) ? '' : 'none';
                });
            });
        }
    </script>
</body>
</html>
"""


# =========================
# Routes
# =========================

@app.route("/", methods=["GET"])
def index():
    workorder = request.args.get("workorder", "").strip()
    status_filter = request.args.get("status_filter", "ALL").strip().upper()

    valid_filters = ["ALL", "SHORT", "NOT ISSUED", "PARTIALLY ISSUED", "FULLY ISSUED"]
    if status_filter not in valid_filters:
        status_filter = "ALL"

    rows = []
    error = None
    summary = {
        "total": 0,
        "not_issued": 0,
        "partial": 0,
        "short": 0,
        "full": 0,
        "total_required": 0,
        "total_issued": 0,
        "total_outstanding": 0,
        "issued_percent": 0,
    }

    if workorder:
        try:
            rows = get_workorder_requirements(workorder, status_filter=status_filter)
            summary = calculate_summary(rows)

        except Exception as ex:
            logger.exception("Failed to load work order requirements")
            error = str(ex)

    return render_template_string(
        HTML,
        workorder=workorder,
        status_filter=status_filter,
        rows=rows,
        summary=summary,
        error=error,
    )


@app.route("/api/workorder/<workorder>", methods=["GET"])
def api_workorder(workorder):
    status_filter = request.args.get("status_filter", "ALL").strip().upper()

    valid_filters = ["ALL", "SHORT", "NOT ISSUED", "PARTIALLY ISSUED", "FULLY ISSUED"]
    if status_filter not in valid_filters:
        status_filter = "ALL"

    try:
        rows = get_workorder_requirements(workorder.strip(), status_filter=status_filter)
        summary = calculate_summary(rows)

        return jsonify({
            "success": True,
            "workorder": workorder,
            "status_filter": status_filter,
            "count": len(rows),
            "summary": summary,
            "data": rows,
        })
    except Exception as ex:
        logger.exception("API error")
        return jsonify({
            "success": False,
            "error": str(ex),
        }), 500


@app.route("/export/workorder/<workorder>", methods=["GET"])
def export_workorder(workorder):
    status_filter = request.args.get("status_filter", "ALL").strip().upper()

    valid_filters = ["ALL", "SHORT", "NOT ISSUED", "PARTIALLY ISSUED", "FULLY ISSUED"]
    if status_filter not in valid_filters:
        status_filter = "ALL"

    try:
        rows = get_workorder_requirements(workorder.strip(), status_filter=status_filter)
        summary = calculate_summary(rows)
        output = create_excel_file(workorder, rows, summary)

        filename = f"workorder_{workorder}_issue_status.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as ex:
        logger.exception("Excel export error")
        return jsonify({
            "success": False,
            "error": str(ex)
        }), 500


@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"})


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", 5000)), debug=True)
